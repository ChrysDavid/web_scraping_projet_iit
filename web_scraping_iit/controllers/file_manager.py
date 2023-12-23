from openpyxl import Workbook, load_workbook
import os

class FolderManager:
    def __init__(self, category):
        self.category = category
        self.create_folder()

    def create_folder(self):
        try:
            category_path = os.path.join('datas', self.category)
            os.makedirs(category_path, exist_ok=True)
            images_path = os.path.join(category_path, 'images')
            os.makedirs(images_path, exist_ok=True)
        except FileExistsError:
            pass

class FileManager:
    def __init__(self, category):
        self.category = category
        self.create_excel_file()

    def create_excel_file(self):
        excel_file = os.path.join('datas', self.category, 'data.xlsx')
        if not os.path.exists(excel_file):
            wb = Workbook()
            ws = wb.active
            headers = ["Nom", "Prix", "Avis", "Description", "En Stock", "Lien de l'Image"]
            ws.append(headers)
            wb.save(excel_file)

    def write_to_excel(self, book_info):
        excel_file = os.path.join('datas', self.category, 'data.xlsx')
        wb = load_workbook(excel_file)
        ws = wb.active
        ws.append(book_info)
        wb.save(excel_file)
